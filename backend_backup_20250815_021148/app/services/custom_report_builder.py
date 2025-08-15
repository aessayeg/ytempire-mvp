"""
Custom Report Builder Service
Enables users to create, customize, and schedule automated reports
"""

import json
import logging
from typing import Dict, List, Any, Optional, Union
from datetime import datetime, timedelta
from enum import Enum
from dataclasses import dataclass, field
import asyncio
import uuid
from pathlib import Path

import pandas as pd
import numpy as np
from jinja2 import Template
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_, or_, func, text

# Optional imports for PDF generation
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

# Optional imports for Excel generation
try:
    import openpyxl
    from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from app.core.config import settings
from app.models.video import Video
from app.models.channel import Channel
from app.models.analytics import Analytics

logger = logging.getLogger(__name__)


class ReportType(Enum):
    """Types of reports available"""
    PERFORMANCE = "performance"
    REVENUE = "revenue"
    CONTENT = "content"
    CHANNEL = "channel"
    COMPETITIVE = "competitive"
    EXECUTIVE = "executive"
    CUSTOM = "custom"


class ReportFormat(Enum):
    """Output formats for reports"""
    PDF = "pdf"
    EXCEL = "excel"
    HTML = "html"
    JSON = "json"
    CSV = "csv"
    POWERPOINT = "powerpoint"


class ReportFrequency(Enum):
    """Report scheduling frequencies"""
    DAILY = "daily"
    WEEKLY = "weekly"
    MONTHLY = "monthly"
    QUARTERLY = "quarterly"
    ON_DEMAND = "on_demand"


@dataclass
class ReportSection:
    """Individual section of a report"""
    title: str
    type: str  # 'text', 'table', 'chart', 'metric'
    content: Any
    order: int
    styling: Dict[str, Any] = field(default_factory=dict)
    filters: Dict[str, Any] = field(default_factory=dict)


@dataclass
class ReportTemplate:
    """Template for report generation"""
    id: str
    name: str
    type: ReportType
    sections: List[ReportSection]
    format: ReportFormat
    frequency: ReportFrequency
    recipients: List[str] = field(default_factory=list)
    filters: Dict[str, Any] = field(default_factory=dict)
    created_at: datetime = field(default_factory=datetime.now)
    last_generated: Optional[datetime] = None


@dataclass
class ReportQuery:
    """SQL query configuration for report data"""
    name: str
    query: str
    parameters: Dict[str, Any] = field(default_factory=dict)
    cache_ttl: int = 3600  # seconds


class CustomReportBuilder:
    """Service for building custom reports"""
    
    def __init__(self):
        self.templates: Dict[str, ReportTemplate] = {}
        self.queries: Dict[str, ReportQuery] = {}
        self.scheduled_reports: Dict[str, Dict[str, Any]] = {}
        self.report_cache: Dict[str, Tuple[Any, datetime]] = {}
        self._initialize_default_templates()
        self._initialize_queries()
    
    def _initialize_default_templates(self):
        """Initialize default report templates"""
        # Executive Summary Report
        exec_template = ReportTemplate(
            id=str(uuid.uuid4()),
            name="Executive Summary",
            type=ReportType.EXECUTIVE,
            format=ReportFormat.PDF,
            frequency=ReportFrequency.WEEKLY,
            sections=[
                ReportSection(
                    title="Key Metrics",
                    type="metric",
                    content={
                        "metrics": ["total_revenue", "total_views", "avg_cpm", "channel_count"]
                    },
                    order=1
                ),
                ReportSection(
                    title="Revenue Trend",
                    type="chart",
                    content={
                        "chart_type": "line",
                        "data_source": "revenue_trend"
                    },
                    order=2
                ),
                ReportSection(
                    title="Top Performing Channels",
                    type="table",
                    content={
                        "data_source": "top_channels",
                        "columns": ["channel_name", "revenue", "views", "videos"]
                    },
                    order=3
                ),
                ReportSection(
                    title="Content Pipeline Status",
                    type="chart",
                    content={
                        "chart_type": "funnel",
                        "data_source": "pipeline_status"
                    },
                    order=4
                )
            ]
        )
        self.templates[exec_template.id] = exec_template
        
        # Performance Report
        perf_template = ReportTemplate(
            id=str(uuid.uuid4()),
            name="Performance Analytics",
            type=ReportType.PERFORMANCE,
            format=ReportFormat.EXCEL,
            frequency=ReportFrequency.DAILY,
            sections=[
                ReportSection(
                    title="Daily Performance",
                    type="table",
                    content={
                        "data_source": "daily_performance",
                        "columns": ["date", "videos_generated", "views", "engagement", "revenue"]
                    },
                    order=1
                ),
                ReportSection(
                    title="Channel Performance",
                    type="table",
                    content={
                        "data_source": "channel_performance",
                        "columns": ["channel", "videos", "views", "avg_duration", "retention"]
                    },
                    order=2
                ),
                ReportSection(
                    title="Content Quality Metrics",
                    type="chart",
                    content={
                        "chart_type": "radar",
                        "data_source": "quality_metrics"
                    },
                    order=3
                )
            ]
        )
        self.templates[perf_template.id] = perf_template
    
    def _initialize_queries(self):
        """Initialize SQL queries for report data"""
        # Revenue trend query
        self.queries['revenue_trend'] = ReportQuery(
            name="Revenue Trend",
            query="""
                SELECT 
                    DATE(created_at) as date,
                    SUM(revenue) as total_revenue,
                    SUM(views) as total_views,
                    AVG(cpm) as avg_cpm
                FROM analytics
                WHERE created_at >= :start_date
                    AND created_at <= :end_date
                GROUP BY DATE(created_at)
                ORDER BY date
            """,
            parameters={"start_date": None, "end_date": None}
        )
        
        # Top channels query
        self.queries['top_channels'] = ReportQuery(
            name="Top Channels",
            query="""
                SELECT 
                    c.name as channel_name,
                    SUM(a.revenue) as revenue,
                    SUM(a.views) as views,
                    COUNT(v.id) as videos
                FROM channels c
                JOIN videos v ON c.id = v.channel_id
                LEFT JOIN analytics a ON v.id = a.video_id
                WHERE a.created_at >= :start_date
                GROUP BY c.id, c.name
                ORDER BY revenue DESC
                LIMIT :limit
            """,
            parameters={"start_date": None, "limit": 10}
        )
        
        # Daily performance query
        self.queries['daily_performance'] = ReportQuery(
            name="Daily Performance",
            query="""
                SELECT 
                    DATE(v.created_at) as date,
                    COUNT(v.id) as videos_generated,
                    SUM(a.views) as views,
                    AVG(a.engagement_rate) as engagement,
                    SUM(a.revenue) as revenue
                FROM videos v
                LEFT JOIN analytics a ON v.id = a.video_id
                WHERE v.created_at >= :start_date
                GROUP BY DATE(v.created_at)
                ORDER BY date DESC
            """,
            parameters={"start_date": None}
        )
    
    async def create_report(
        self,
        template_id: str,
        db: AsyncSession,
        custom_filters: Optional[Dict[str, Any]] = None,
        format_override: Optional[ReportFormat] = None
    ) -> Dict[str, Any]:
        """Create a report based on template"""
        if template_id not in self.templates:
            raise ValueError(f"Template {template_id} not found")
        
        template = self.templates[template_id]
        report_format = format_override or template.format
        
        # Merge filters
        filters = {**template.filters, **(custom_filters or {})}
        
        # Generate report data
        report_data = {
            'id': str(uuid.uuid4()),
            'name': template.name,
            'type': template.type.value,
            'generated_at': datetime.now().isoformat(),
            'filters': filters,
            'sections': []
        }
        
        # Process each section
        for section in sorted(template.sections, key=lambda x: x.order):
            section_data = await self._process_section(section, db, filters)
            report_data['sections'].append(section_data)
        
        # Generate output in requested format
        output = await self._generate_output(report_data, report_format)
        
        # Update last generated timestamp
        template.last_generated = datetime.now()
        
        return {
            'report_id': report_data['id'],
            'format': report_format.value,
            'data': output,
            'metadata': {
                'template': template.name,
                'generated_at': report_data['generated_at'],
                'sections': len(report_data['sections'])
            }
        }
    
    async def _process_section(
        self,
        section: ReportSection,
        db: AsyncSession,
        filters: Dict[str, Any]
    ) -> Dict[str, Any]:
        """Process individual report section"""
        section_data = {
            'title': section.title,
            'type': section.type,
            'content': None
        }
        
        if section.type == 'metric':
            section_data['content'] = await self._process_metrics(
                section.content['metrics'], db, filters
            )
        elif section.type == 'table':
            section_data['content'] = await self._process_table(
                section.content['data_source'], 
                section.content.get('columns'), 
                db, 
                filters
            )
        elif section.type == 'chart':
            section_data['content'] = await self._process_chart(
                section.content['chart_type'],
                section.content['data_source'],
                db,
                filters
            )
        elif section.type == 'text':
            section_data['content'] = await self._process_text(
                section.content, db, filters
            )
        
        return section_data
    
    async def _process_metrics(
        self,
        metrics: List[str],
        db: AsyncSession,
        filters: Dict[str, Any]
    ) -> Dict[str, Any]:
        """Process metric calculations"""
        result = {}
        
        for metric in metrics:
            if metric == 'total_revenue':
                # Simulate metric calculation
                result[metric] = {
                    'value': np.random.uniform(10000, 50000),
                    'change': np.random.uniform(-10, 20),
                    'period': 'last_30_days'
                }
            elif metric == 'total_views':
                result[metric] = {
                    'value': np.random.randint(100000, 1000000),
                    'change': np.random.uniform(-5, 15),
                    'period': 'last_30_days'
                }
            elif metric == 'avg_cpm':
                result[metric] = {
                    'value': np.random.uniform(2, 8),
                    'change': np.random.uniform(-2, 3),
                    'period': 'last_30_days'
                }
            elif metric == 'channel_count':
                result[metric] = {
                    'value': np.random.randint(10, 50),
                    'change': np.random.randint(0, 5),
                    'period': 'total'
                }
        
        return result
    
    async def _process_table(
        self,
        data_source: str,
        columns: Optional[List[str]],
        db: AsyncSession,
        filters: Dict[str, Any]
    ) -> Dict[str, Any]:
        """Process table data"""
        # Fetch data based on data source
        if data_source in self.queries:
            query = self.queries[data_source]
            # Apply filters to query parameters
            params = {**query.parameters}
            params.update(filters)
            
            # Simulate data fetch (replace with actual DB query)
            data = self._simulate_table_data(data_source, columns)
        else:
            data = []
        
        return {
            'columns': columns or [],
            'rows': data,
            'total': len(data)
        }
    
    def _simulate_table_data(
        self,
        data_source: str,
        columns: Optional[List[str]]
    ) -> List[Dict[str, Any]]:
        """Simulate table data for demonstration"""
        if data_source == 'top_channels':
            data = []
            for i in range(10):
                data.append({
                    'channel_name': f'Channel_{i+1}',
                    'revenue': np.random.uniform(1000, 10000),
                    'views': np.random.randint(10000, 100000),
                    'videos': np.random.randint(10, 100)
                })
            return data
        elif data_source == 'daily_performance':
            data = []
            for i in range(30):
                date = datetime.now() - timedelta(days=i)
                data.append({
                    'date': date.strftime('%Y-%m-%d'),
                    'videos_generated': np.random.randint(5, 20),
                    'views': np.random.randint(5000, 50000),
                    'engagement': np.random.uniform(0.1, 0.5),
                    'revenue': np.random.uniform(100, 1000)
                })
            return data
        return []
    
    async def _process_chart(
        self,
        chart_type: str,
        data_source: str,
        db: AsyncSession,
        filters: Dict[str, Any]
    ) -> Dict[str, Any]:
        """Process chart data"""
        # Fetch data for chart
        if data_source == 'revenue_trend':
            dates = pd.date_range(end=datetime.now(), periods=30, freq='D')
            data = {
                'x': [d.strftime('%Y-%m-%d') for d in dates],
                'y': [np.random.uniform(500, 2000) for _ in range(30)],
                'type': chart_type
            }
        elif data_source == 'pipeline_status':
            data = {
                'labels': ['Trending', 'Script', 'Voice', 'Video', 'Published'],
                'values': [100, 85, 70, 65, 60],
                'type': 'funnel'
            }
        elif data_source == 'quality_metrics':
            data = {
                'categories': ['Script Quality', 'Voice Quality', 'Video Quality', 'SEO Score', 'Engagement'],
                'values': [85, 90, 88, 92, 78],
                'type': 'radar'
            }
        else:
            data = {'type': chart_type}
        
        return data
    
    async def _process_text(
        self,
        content: Any,
        db: AsyncSession,
        filters: Dict[str, Any]
    ) -> str:
        """Process text content with template variables"""
        if isinstance(content, str):
            # Process template variables
            template = Template(content)
            return template.render(**filters)
        return str(content)
    
    async def _generate_output(
        self,
        report_data: Dict[str, Any],
        format: ReportFormat
    ) -> Any:
        """Generate report output in specified format"""
        if format == ReportFormat.JSON:
            return json.dumps(report_data, indent=2, default=str)
        elif format == ReportFormat.HTML:
            return self._generate_html_report(report_data)
        elif format == ReportFormat.PDF and REPORTLAB_AVAILABLE:
            return self._generate_pdf_report(report_data)
        elif format == ReportFormat.EXCEL and OPENPYXL_AVAILABLE:
            return self._generate_excel_report(report_data)
        elif format == ReportFormat.CSV:
            return self._generate_csv_report(report_data)
        else:
            # Fallback to JSON
            return json.dumps(report_data, indent=2, default=str)
    
    def _generate_html_report(self, report_data: Dict[str, Any]) -> str:
        """Generate HTML report"""
        html_template = """
        <!DOCTYPE html>
        <html>
        <head>
            <title>{title}</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                h1 {{ color: #333; }}
                h2 {{ color: #666; border-bottom: 2px solid #ddd; padding-bottom: 5px; }}
                table {{ border-collapse: collapse; width: 100%; margin: 20px 0; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                th {{ background-color: #f2f2f2; }}
                .metric {{ display: inline-block; margin: 10px; padding: 15px; 
                          background: #f9f9f9; border-radius: 5px; }}
                .metric-value {{ font-size: 24px; font-weight: bold; color: #2196F3; }}
                .metric-label {{ color: #666; margin-top: 5px; }}
                .chart {{ margin: 20px 0; padding: 20px; background: #f9f9f9; }}
            </style>
        </head>
        <body>
            <h1>{title}</h1>
            <p>Generated: {generated_at}</p>
            {sections}
        </body>
        </html>
        """
        
        sections_html = ""
        for section in report_data['sections']:
            sections_html += f"<h2>{section['title']}</h2>"
            
            if section['type'] == 'metric':
                sections_html += '<div class="metrics">'
                for metric_name, metric_data in section['content'].items():
                    sections_html += f"""
                    <div class="metric">
                        <div class="metric-value">{metric_data['value']:.2f}</div>
                        <div class="metric-label">{metric_name.replace('_', ' ').title()}</div>
                    </div>
                    """
                sections_html += '</div>'
            
            elif section['type'] == 'table':
                sections_html += '<table>'
                sections_html += '<tr>'
                for col in section['content']['columns']:
                    sections_html += f'<th>{col}</th>'
                sections_html += '</tr>'
                
                for row in section['content']['rows']:
                    sections_html += '<tr>'
                    for col in section['content']['columns']:
                        sections_html += f'<td>{row.get(col, "")}</td>'
                    sections_html += '</tr>'
                sections_html += '</table>'
            
            elif section['type'] == 'chart':
                sections_html += f'<div class="chart">Chart: {section["content"].get("type", "")}</div>'
        
        return html_template.format(
            title=report_data['name'],
            generated_at=report_data['generated_at'],
            sections=sections_html
        )
    
    def _generate_pdf_report(self, report_data: Dict[str, Any]) -> bytes:
        """Generate PDF report using ReportLab"""
        if not REPORTLAB_AVAILABLE:
            return self._generate_html_report(report_data).encode()
        
        # Create PDF buffer
        from io import BytesIO
        buffer = BytesIO()
        
        # Create PDF document
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#333333')
        )
        story.append(Paragraph(report_data['name'], title_style))
        story.append(Spacer(1, 12))
        
        # Generated timestamp
        story.append(Paragraph(f"Generated: {report_data['generated_at']}", styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Process sections
        for section in report_data['sections']:
            # Section title
            story.append(Paragraph(section['title'], styles['Heading2']))
            story.append(Spacer(1, 12))
            
            if section['type'] == 'table' and section['content']['rows']:
                # Create table
                table_data = [section['content']['columns']]
                for row in section['content']['rows']:
                    table_data.append([row.get(col, '') for col in section['content']['columns']])
                
                t = Table(table_data)
                t.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                story.append(t)
            
            story.append(Spacer(1, 20))
        
        # Build PDF
        doc.build(story)
        pdf_data = buffer.getvalue()
        buffer.close()
        
        return pdf_data
    
    def _generate_excel_report(self, report_data: Dict[str, Any]) -> bytes:
        """Generate Excel report using openpyxl"""
        if not OPENPYXL_AVAILABLE:
            return self._generate_csv_report(report_data).encode()
        
        from io import BytesIO
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Create summary sheet
        ws = wb.active
        ws.title = "Summary"
        
        # Add title
        ws['A1'] = report_data['name']
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = f"Generated: {report_data['generated_at']}"
        
        current_row = 4
        
        # Process sections
        for section in report_data['sections']:
            # Section title
            ws.cell(row=current_row, column=1, value=section['title'])
            ws.cell(row=current_row, column=1).font = Font(size=12, bold=True)
            current_row += 2
            
            if section['type'] == 'table' and section['content']['rows']:
                # Add headers
                for col_idx, col_name in enumerate(section['content']['columns'], 1):
                    ws.cell(row=current_row, column=col_idx, value=col_name)
                    ws.cell(row=current_row, column=col_idx).font = Font(bold=True)
                    ws.cell(row=current_row, column=col_idx).fill = PatternFill(
                        start_color="DDDDDD", end_color="DDDDDD", fill_type="solid"
                    )
                
                current_row += 1
                
                # Add data rows
                for row_data in section['content']['rows']:
                    for col_idx, col_name in enumerate(section['content']['columns'], 1):
                        ws.cell(row=current_row, column=col_idx, value=row_data.get(col_name, ''))
                    current_row += 1
                
                current_row += 2
            
            elif section['type'] == 'metric':
                for metric_name, metric_data in section['content'].items():
                    ws.cell(row=current_row, column=1, value=metric_name.replace('_', ' ').title())
                    ws.cell(row=current_row, column=2, value=metric_data['value'])
                    ws.cell(row=current_row, column=3, value=f"Change: {metric_data['change']:.1f}%")
                    current_row += 1
                current_row += 2
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        excel_data = buffer.getvalue()
        buffer.close()
        
        return excel_data
    
    def _generate_csv_report(self, report_data: Dict[str, Any]) -> str:
        """Generate CSV report"""
        csv_lines = []
        csv_lines.append(f"Report: {report_data['name']}")
        csv_lines.append(f"Generated: {report_data['generated_at']}")
        csv_lines.append("")
        
        for section in report_data['sections']:
            csv_lines.append(f"Section: {section['title']}")
            
            if section['type'] == 'table' and section['content']['rows']:
                # Add headers
                csv_lines.append(','.join(section['content']['columns']))
                
                # Add data rows
                for row in section['content']['rows']:
                    row_values = [str(row.get(col, '')) for col in section['content']['columns']]
                    csv_lines.append(','.join(row_values))
            
            csv_lines.append("")
        
        return '\n'.join(csv_lines)
    
    def schedule_report(
        self,
        template_id: str,
        frequency: ReportFrequency,
        recipients: List[str],
        time_of_day: Optional[str] = None
    ) -> str:
        """Schedule automatic report generation"""
        schedule_id = str(uuid.uuid4())
        
        self.scheduled_reports[schedule_id] = {
            'template_id': template_id,
            'frequency': frequency,
            'recipients': recipients,
            'time_of_day': time_of_day or '09:00',
            'next_run': self._calculate_next_run(frequency, time_of_day),
            'active': True
        }
        
        logger.info(f"Scheduled report {schedule_id} for template {template_id}")
        return schedule_id
    
    def _calculate_next_run(
        self,
        frequency: ReportFrequency,
        time_of_day: Optional[str] = None
    ) -> datetime:
        """Calculate next run time for scheduled report"""
        now = datetime.now()
        
        if frequency == ReportFrequency.DAILY:
            next_run = now + timedelta(days=1)
        elif frequency == ReportFrequency.WEEKLY:
            next_run = now + timedelta(weeks=1)
        elif frequency == ReportFrequency.MONTHLY:
            next_run = now + timedelta(days=30)
        elif frequency == ReportFrequency.QUARTERLY:
            next_run = now + timedelta(days=90)
        else:
            next_run = now
        
        # Set specific time if provided
        if time_of_day:
            hour, minute = map(int, time_of_day.split(':'))
            next_run = next_run.replace(hour=hour, minute=minute, second=0, microsecond=0)
        
        return next_run
    
    def create_custom_template(
        self,
        name: str,
        sections: List[Dict[str, Any]],
        format: ReportFormat = ReportFormat.PDF,
        frequency: ReportFrequency = ReportFrequency.ON_DEMAND
    ) -> str:
        """Create custom report template"""
        template_id = str(uuid.uuid4())
        
        # Convert section dicts to ReportSection objects
        report_sections = []
        for idx, section_data in enumerate(sections):
            report_sections.append(ReportSection(
                title=section_data.get('title', f'Section {idx+1}'),
                type=section_data.get('type', 'text'),
                content=section_data.get('content', {}),
                order=section_data.get('order', idx),
                styling=section_data.get('styling', {}),
                filters=section_data.get('filters', {})
            ))
        
        template = ReportTemplate(
            id=template_id,
            name=name,
            type=ReportType.CUSTOM,
            sections=report_sections,
            format=format,
            frequency=frequency
        )
        
        self.templates[template_id] = template
        logger.info(f"Created custom template: {name} (ID: {template_id})")
        
        return template_id
    
    def get_templates(self) -> List[Dict[str, Any]]:
        """Get list of available report templates"""
        return [
            {
                'id': template_id,
                'name': template.name,
                'type': template.type.value,
                'format': template.format.value,
                'frequency': template.frequency.value,
                'sections': len(template.sections),
                'last_generated': template.last_generated.isoformat() if template.last_generated else None
            }
            for template_id, template in self.templates.items()
        ]
    
    def get_scheduled_reports(self) -> List[Dict[str, Any]]:
        """Get list of scheduled reports"""
        return [
            {
                'id': schedule_id,
                'template_id': schedule['template_id'],
                'frequency': schedule['frequency'].value if isinstance(schedule['frequency'], ReportFrequency) else schedule['frequency'],
                'recipients': schedule['recipients'],
                'next_run': schedule['next_run'].isoformat() if isinstance(schedule['next_run'], datetime) else schedule['next_run'],
                'active': schedule['active']
            }
            for schedule_id, schedule in self.scheduled_reports.items()
        ]


# Singleton instance
custom_report_builder = CustomReportBuilder()